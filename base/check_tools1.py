import time
import string

from openpyxl.utils import get_column_letter

from read_all_files import find_file
from trans_reading import read_xlsx_file, rows, read_csv_file
from deepdiff import DeepDiff
import re
import shutil
import os
from time import sleep
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from database_tools import execute_sql
from translate import translate_text


def start_check(channel):
    global language1, language2
    files = find_file(f"../data/{channel}_data", include_str="language", filter_strs=["~"])
    # mac排序与win相反
    fil = files[:-3:-1]
    # fil = files
    if channel == 'server':
        language1 = read_csv_file(fil[0])
        language2 = read_csv_file(fil[1])
        check_tools(channel)
    else:
        language1 = read_xlsx_file(fil[0])
        language2 = read_xlsx_file(fil[1])
        check_tools(channel)


# 检测
def check_tools(channel):
    max1 = rows(language1)
    max2 = rows(language2)
    if max1 == max2:
        # logger.info(F"本次行数相同,共计key{max1 - 1}条")
        datas_key = different_key()
        if len(datas_key) > 0:
            # logger.info(f"本次修改了key，{datas_key}")
            datas = different_data(language1)
            msg = [f"本次修改了key，{datas_key}"]
            msg2 = []
            data = ""
            generate_xlsx(file=language1, file_list=datas, msg=msg, channel=channel, msg2=msg2, datas=data)
        else:
            # logger.info("本次内容未新增key,下面进行内容检查")
            dif_msg = different_msg()[0]
            if len(dif_msg) > 0:
                msg = [
                    f"本次检测共有{len(dif_msg)}条的值出现变化,修改后的详情见下方！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！"]
                msg2 = []
                data = ""
                print(msg)
                generate_xlsx(file=language1, file_list=dif_msg, msg=msg, channel=channel, msg2=msg2, datas=data)
            else:
                # logger.info(f"本次未修改KEY，也未对值进行修改")
                print(f"本次未修改KEY，也未对值进行修改")
    elif max1 < max2:
        rol = different_row_number()
        rol = [i + 2 for i in rol]
        msg = [f"本次多语言在{rol}行减少,共减少{max2 - max1}条"]
        # datas_key = different_key()
        datas = different_data(language2)
        # logger.info(f"第{rol}行减少key有{datas_key}")
        msg2 = []
        data = ""
        generate_xlsx(file=language2, file_list=datas, msg=msg, channel=channel, msg2=msg2, datas=data)
    elif max1 > max2:
        # datas_key = different_key()
        rol = different_row_number()
        rol = [i + 2 for i in rol]
        msg = [f"本次多语言在{rol}行新增,共新增{max1 - max2}条"]
        datas1 = different_data(language1)
        translate_date = translated_datas(datas1)
        if channel == 'server':
            generate_xlsx(file=language1, file_list=[datas1, ""], msg=msg, channel=channel, msg2=[''], datas="")
            # execute_sql(channel_id=channel_num(channel), newly_quantity=max1 - max2,
            #             modify_quantity=0, quantity=max1)
        else:
            datas2 = add_change_diff(language1)
            datas = [translate_date, datas2[0]]
            # logger.info(f"第{rol}增加key{datas_key}")
            if len(datas2[0]) > 0:
                msg2 = [
                    f"本次检测共有{len(datas2[0])}条多语言的值出现变化,修改后的详情见下方！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！"]
            else:
                msg2 = [f"本次只有新增，没有修改多语言"]
            generate_xlsx(file=language1, file_list=datas, msg=msg, channel=channel, msg2=msg2, datas=datas2)
            # execute_sql(channel_id=channel_num(channel), newly_quantity=max1 - max2,
            #             modify_quantity=len(datas2[0]), quantity=max1)

        # 获取差异key 与行数


def translated_datas(original_list):
    translated_list = []

    for sublist in original_list:
        translated_sublist = []
        for text in sublist:
            if text is None or str(text).lower() == 'nan' or '_' in text:
                translated_sublist.append(text)
                continue
            else:
                translated_sublist.append(f"{text}:::{translate_text(text)}")
        translated_list.append(translated_sublist)
    return translated_list


def different_key():
    old_file = language2.iloc[:, 0].tolist()
    new_file = language1.iloc[:, 0].tolist()
    diff = DeepDiff(old_file, new_file)
    return diff


# 列数相同时，对比所有差异数据
def different_msg():
    diff_msg = []
    diff = []
    old_file = language2.values.tolist()
    new_file = language1.values.tolist()
    for i in range(len(new_file) - 1, -1, -1):
        if new_file[i] != old_file[i]:
            diff_msg.append(new_file[i])
            a = DeepDiff(old_file[i], new_file[i])
            diff.append(a)
    return diff_msg, diff


def channel_num(channel):
    """
    :param channel: 端名称
    :return: 返回对应的id
    """
    channel_dict = {"android": 1, "ios": 2, 'server': 3, 'unity': 4}
    channel_number = channel_dict[channel]
    return channel_number


# 既有变化也有修改
def add_change_diff(language):
    diff_msg = []
    diff = []
    diff_num = different_row_number()
    diff_data = different_data(language)
    old_file = language2.values.tolist()
    new_file = language.values.tolist()
    for data in range(len(diff_num)):
        old_file.insert(diff_num[data], diff_data[data])
    for i in range(len(new_file) - 1, -1, -1):
        if new_file[i] != old_file[i]:
            diff_msg.append(new_file[i])
            a = DeepDiff(old_file[i], new_file[i])
            diff.append(a)
    return diff_msg, diff


def sort_list(file_list):
    """
    :param file_list: 传入列表
    :return: 返回排序后的列表
    """
    new_file_list = sorted(file_list, key=lambda x: x[0].lower())
    original_case_map = {word.lower(): word for word in file_list}
    sorted_list = [[original_case_map[word.lower()] for word in sublist] for sublist in new_file_list]
    return sorted_list


# 获取差异行
def different_row_number():
    key_rol = []
    diff = different_key()
    diff_keys = diff.keys()
    for key in diff_keys:
        dif_values = diff[key]
        for keys in dif_values:
            numbers_list = re.findall(r'\d+', keys)
            numbers = "".join(numbers_list)
            key_rol.append(int(numbers))
    return key_rol


# 获取每行差异数据
def different_data(file_name):
    data_list = []
    dif_num = different_row_number()
    for data_num in dif_num:
        data_list.append(file_name.loc[data_num].to_list())
    return data_list


# 写入表格文件
def generate_xlsx(file, file_list, msg, msg2, channel, datas):
    times = time.strftime('%Y年%m月%d日 %H点-%M分-%S秒', time.localtime(time.time()))
    new_name = f"../result/{times}--{channel}--language_test.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    set_column_width(sheet, channel)
    sheet.append(msg)
    head = change_head(file)
    sheet.append(head)
    time.sleep(2)
    for file in file_list:
        for i in file:
            sheet.append(i)
        if msg2:
            sheet.append(msg2)
            msg2 = False
    insert_edit_cols(sheet)
    if datas != "":
        get_line_value(datas[1], sheet)
    workbook.save(new_name)


def set_column_width(sheet, channel):
    """
    :param sheet: 继承sheet方法
    :param channel: 传入端名
    :return:设置列宽
    """
    if channel == 'android':
        for i in range(3, 27):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = 30
    elif channel == 'ios':
        for i in range(3, 24):
            if i == 6 or i == 20:
                pass
            else:
                column_letter = get_column_letter(i)
                sheet.column_dimensions[column_letter].width = 30


def get_head(file):
    head = file.keys().tolist()
    return head


def change_head(file):
    """
    :param file:传入文件
    :return: 返回修改后的表头
    """
    head_list = get_head(file)
    lan_dic = language_dic()
    for i in range(len(head_list)):
        if head_list[i] in lan_dic:
            head_list[i] = lan_dic[head_list[i]]
    return head_list


def language_dic():
    lan_dic = {"ar": "ar：阿拉伯语", 'en': "en：英语🇬🇧", 'bn': "bn-IN：孟加拉语-印度🇧🇩", 'bn-IN': "bn-IN：孟加拉语-印度🇧🇩",
               'cs': "cs：捷克语🇨🇿", 'de': "de：德语🇩🇪", 'es': "se：西班牙语🇪🇸", 'fr': "fr：法语🇫🇷", 'id': "id：印尼语🇮🇩",
               'in': "in：印尼语🇮🇩", "it": "it：意大利语🇮🇹", 'ja': "ja：日语🇯🇵", 'ko': "ko：韩语🇰🇷", 'ms': "ms：马来语🇲🇾",
               'pt-rBR': "pt-BR：葡萄牙语🇵🇹", 'pt-BR': "pt-BR：葡萄牙语🇵🇹", 'ru': "ru：俄语🇷🇺（老）",
               'ru-rRU': "ru-rRU：俄语🇷🇺", "ru-RU": "ru-rRU：俄语🇷🇺", 'sr': "sr：塞尔维亚语🇷🇸", 'th': "th：泰语🇹🇭",
               "tr": "tr：土耳其语🇹🇷（老）", 'tr-rTR': "tr-rTR：土耳其语🇹🇷（新）", 'tr-TR': "tr-TR：土耳其语🇹🇷",
               'ur-PK': "ur-rPK：ur-PK：乌尔都语🇵🇰", "ur-rPK": "ur-rPK：乌尔都语🇵🇰", 'vi': 'vi：越南语🇻🇳',
               'zh-rCN': "zh-rCN：中文🇨🇳", "zh-CN": "zh-CN：中文🇨🇳", 'zh-rTW': "zh-rTW：繁体中文🇨🇳",
               "zh-Hant": "zh-Hant:繁体中文🇨🇳"}
    return lan_dic


# 文件名修改
def change_filename(client):
    Dpath = os.path.expanduser(r'~\Downloads')
    import time
    times = time.strftime('%Y年%m月%d日 %H点-%M分-%S秒', time.localtime(time.time()))
    new_name = f"{times}language_{client}.xlsx"
    en_file = find_file(Dpath, include_str='en', filter_strs=[".~"])
    os.rename(en_file[0], new_name)


# 文件移动
def move_file(client):
    change_filename(client)
    data_path = f"../data/{client}_data"
    en_file = find_file(f'..//case', include_str=f'language_{client}', filter_strs=[".~"])
    shutil.move(en_file[0], data_path)
    sleep(3)


# 插入编辑列
def insert_edit_cols(sheet):
    sheet.insert_cols(idx=1, amount=2)
    sheet["A2"] = "完成备注"
    sheet['B2'] = "编号"


# 插入变化值
def get_line_value(values, sheet):
    max1 = rows(language1)
    max2 = rows(language2)
    line = max1 - max2 + 4
    for i in range(len(values)):
        for k, v in values[i].items():
            for k1, v1 in v.items():
                line1 = change_int(k1) + 3
                color_fill(sheet, line, line1)
                comment = Comment(str(v1), "hhhh")
                letters = number_to_letter(line1)
                str_num = str(line)
                nl = letters + str_num
                sheet[nl].comment = comment
        line += 1


# 转化数字
def change_int(a):
    result = ""
    for char in a:
        if char.isdigit():
            result += char
    return int(result)


def number_to_letter(number):
    if 1 <= number <= 26:
        return string.ascii_uppercase[number - 1]
    elif number > 26:
        return "A" + string.ascii_uppercase[number - 26 - 1]


# 颜色填充
def color_fill(sheet, row, column):
    fill = PatternFill('solid', fgColor='f8c600')
    sheet.cell(row, column).fill = fill


def del_file():
    """
    :return:清空result下文件
    """
    import os
    folder_path = "../result"
    file_list = os.listdir(folder_path)
    for file in file_list:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            os.remove(file_path)

    print("已成功删除文件夹下的所有文件")


start_check('ios')
