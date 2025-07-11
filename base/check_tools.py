import time
import string
from openpyxl.utils import get_column_letter
from base.read_all_files import find_file
from base.trans_reading import read_xlsx_file, rows, read_csv_file
from deepdiff import DeepDiff
import re
import shutil
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from base.database_tools import execute_sql
from base.translate import translate_text
from base.baidu_ai import main
import pandas as pd


def start_check(channel):
    """
    :param channel:传入对应的端
    :return:
    """
    global language1, language2
    files = find_file(fr"{absolute_path('data')}/{channel}_data", include_str="language",
                      filter_strs=["~"])
    fil = files[:-3:-1]
    if channel == 'server':
        sort_excel_by_first_column_desc(fil[0])
        language1 = read_xlsx_file(fil[0])
        language2 = read_xlsx_file(r"D:\project\starx_project\translate\data\server_data\2025年04月07日 "
                                   r"14点-20分language_server.xlsx")
        check_tools(channel)
    else:
        sort_excel_by_first_column_desc(fil[0])
        sort_excel_by_first_column_desc(fil[1])
        language1 = read_xlsx_file(fil[0])
        language2 = read_xlsx_file(fil[1])
        check_tools(channel)

# 检测
def check_tools(channel):
    """
    :param channel: 选择对应端
    :return: 返回差异内容
    """
    max1 = rows(language1)
    max2 = rows(language2)
    if max1 == max2:
        # logger.info(F"本次行数相同,共计key{max1 - 1}条")
        datas_key = different_key()
        if len(datas_key) > 0:
            # logger.info(f"本次修改了key，{datas_key}")
            datas = different_data(language1)
            msg = [f"本次修改了key，{datas_key}"]
            msg2 = []
            data = ""
            generate_xlsx(num=0, file=language1, file_list=datas, msg=msg, channel=channel, msg2=msg2, datas=data)
        else:
            # logger.info("本次内容未新增key,下面进行内容检查")
            dif_msg = different_msg()
            if len(dif_msg[0]) > 0:
                msg = [
                    f"本次检测共有{len(dif_msg[0])}条的值出现变化,修改后的详情见下方！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！"]
                msg2 = []
                data = ""
                generate_xlsx(num=0, file=language1, file_list=dif_msg[0], msg=msg, channel=channel, msg2=msg2,
                              datas=data)
                execute_sql(channel_id=channel_num(channel), newly_quantity=max1 - max2,
                            modify_quantity=len(dif_msg[0]), quantity=max1)
            else:
                # logger.info(f"本次未修改KEY，也未对值进行修改")
                print(f"本次未修改KEY，也未对值进行修改")
    elif max1 < max2:
        rol = different_row_number()
        rol = [i + 2 for i in rol]
        msg = [f"本次多语言在{rol}行减少,共减少{max2 - max1}条"]
        # datas_key = different_key()
        datas = different_data(language2)
        # logger.info(f"第{rol}行减少key有{datas_key}")
        msg2 = []
        data = ""
        generate_xlsx(num=0, file=language2, file_list=datas, msg=msg, channel=channel, msg2=msg2, datas=data)
    elif max1 > max2:
        # datas_key = different_key()
        rol = different_row_number()
        rol = [i + 2 for i in rol]
        msg = [f"本次多语言在{rol}行新增,共新增{max1 - max2}条"]
        datas1 = different_data(language1)
        if channel == 'server':
            generate_xlsx(num=0, file=language1, file_list=[datas1, ""], msg=msg, channel=channel, msg2=[''], datas="")
            execute_sql(channel_id=channel_num(channel), newly_quantity=max1 - max2,
                        modify_quantity=0, quantity=max1)
        else:
            translate_date = translated_datas(datas1, channel)
            datas2 = add_change_diff(language1)
            datas = [translate_date, datas2[0]]
            # logger.info(f"第{rol}增加key{datas_key}")
            if len(datas2[0]) > 0:
                msg2 = [
                    f"本次检测共有{len(datas2[0])}条多语言的值出现变化,修改后的详情见下方！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！"]
            else:
                msg2 = [f"本次只有新增，没有修改多语言"]
            generate_xlsx(num=max1 - max2, file=language1, file_list=datas, msg=msg, channel=channel, msg2=msg2,
                          datas=datas2)
            execute_sql(channel_id=channel_num(channel), newly_quantity=max1 - max2,
                        modify_quantity=len(datas2[0]), quantity=max1, method="insert")


def translated_datas(original_list, channel):
    """
    翻译对应语言
    """
    ios_language_list = ['en', 'zh-cn', 'de', 'es', 'fr', 'auto', 'ar', 'bn', 'id', 'it', 'ja', 'ko', 'ms', 'pt', 'ru',
                         'th', 'tr', 'vi']
    android_language_list = ['en', 'zh-cn', 'de', 'es', 'fr', 'auto', 'ar', 'bn', 'id', 'it', 'ja', 'ko', 'ms', 'pt',
                             'ru', 'th', 'tr', 'ur', 'vi']
    server_language_list = ['en', 'ar',  'bn', 'de',  'es', 'fr',  'id',
                            'it', 'ja',  'ko',  'ms', 'pt', 'ru',
                            'th', 'tr', 'ur', 'vi', 'zh-cn', 'auto', 'auto']
    flutter_language_list = ['en', 'ar', 'bn', 'cs', 'de', 'es', 'fr', 'id', 'it', 'ja', 'ko', 'ms', 'pt', 'ru', 'sr',
                             'th', 'tr', 'ur',
                             'vi', 'zh-cn', 'auto', 'auto']
    if channel == 'ios':
        return translated_datas_start(original_list, ios_language_list)
    elif channel == 'android':
        return translated_datas_start(original_list, android_language_list)
    elif channel == 'server':
        return translated_datas_start(original_list, server_language_list)
    else:
        return translated_datas_start(original_list, flutter_language_list)


def translated_datas_start(original_list, language):
    """
    开启翻译
    """
    translated_list = []
    for sublist in original_list:
        translated_sublist = []
        translate = []
        t = 0
        translated_sublist.append(sublist[0])
        sublist.remove(sublist[0])
        for text in sublist:
            if text is None or str(text).lower() == 'nan':
                translated_sublist.append(text)
                t += 1
            else:
                hans = translate_text(text, src=language[t])
                translated_sublist.append(f"“{text}”翻译:{hans}")
                t += 1
                print(translated_sublist)
                translate.append(hans)
                translate.append(hans)
        trans_data = main(translate)
        if "不相近" in trans_data or "不" in trans_data:
            translated_sublist.append(trans_data)
        else:
            translated_sublist.append('该多语言文案翻译的意思相近')
        translated_list.append(translated_sublist)
    return translated_list


def channel_num(channel):
    """
    :param channel: 端名称
    :return: 返回对应的id
    """
    channel_dict = {"android": 1, "ios": 2, 'server': 3, 'unity': 4, 'flutter': 5, 'cocos': 6}
    channel_number = channel_dict[channel]
    return channel_number


def different_key():
    """
    :return: 获取差异key 与行数
    """
    old_file = language2.iloc[:, 0].tolist()
    new_file = language1.iloc[:, 0].tolist()
    diff = DeepDiff(old_file, new_file)
    return diff


def different_msg():
    """
    :return: 没有新增时，对比所有差异数据
    """
    diff_msg = []
    diff = []
    old_file = language2.values.tolist()
    new_file = language1.values.tolist()
    for i in range(len(new_file) - 1, -1, -1):
        if new_file[i] != old_file[i]:
            diff_msg.append(new_file[i])
            a = DeepDiff(old_file[i], new_file[i])
            diff.append(a)
    return diff_msg, diff


# 既有变化也有修改
def add_change_diff(language):
    """
    :param language: 传入语言列表
    :return: 返回变化内容与变化
    """
    diff_msg = []
    diff = []
    diff_num = different_row_number()
    diff_data = different_data(language)
    old_file = language2.values.tolist()
    new_file = language.values.tolist()
    for data in range(len(diff_num)):
        old_file.insert(diff_num[data], diff_data[data])
    for i in range(len(new_file) - 1, -1, -1):
        if new_file[i] != old_file[i]:
            diff_msg.append(new_file[i])
            a = DeepDiff(old_file[i], new_file[i])
            diff.append(a)
    return diff_msg, diff


def different_row_number():
    """
    :return: 返回差异行
    """
    key_rol = []
    diff = different_key()
    diff_keys = diff.keys()
    for key in diff_keys:
        dif_values = diff[key]
        for keys in dif_values:
            numbers_list = re.findall(r'\d+', keys)
            numbers = "".join(numbers_list)
            key_rol.append(int(numbers))
    return key_rol


def different_data(file_name):
    """
    :param file_name: 传入文件名
    :return:  获取每行差异数据
    """
    data_list = []
    dif_num = different_row_number()
    for data_num in dif_num:
        data_list.append(file_name.loc[data_num].to_list())
    return data_list


def generate_xlsx(num, file, file_list, msg, msg2, channel, datas):
    """
    :param num: 列数
    :param file:文件名
    :param file_list:
    :param msg: 写入新增值
    :param msg2:写入变换值
    :param channel:创建文件名的端名
    :param datas:变化的内容
    :return: 写入表格文件
    """
    times = time.strftime('%Y年%m月%d日 %H点-%M分', time.localtime(time.time()))
    new_name = fr"{absolute_path('result')}\{times}--{channel}--language_test.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    set_column_width(sheet, channel)
    sheet.append(msg)
    head = change_head(file)
    sheet.append(head)
    for file in file_list:
        for i in file:
            sheet.append(i)
        if msg2:
            sheet.append(msg2)
            msg2 = False
    insert_edit_cols(sheet)
    if datas != "":
        get_line_value(datas[1], sheet)
    # del_cols(channel, sheet)
    if num == 0:
        workbook.save(new_name)
    else:
        variable_check(sheet)
        workbook.save(new_name)


def variable_check(sheet):
    """
    :param sheet: 传入指定表
    :return: 返回表格上色
    """
    for row_num, row in enumerate(sheet.iter_rows(min_col=4, max_col=4, values_only=False), start=1):
        cell = row[0]
        if "{" in str(cell.value) or "%" in str(cell.value):
            print(cell.value)
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            c_cell = sheet.cell(row=row_num, column=3)
            c_cell.fill = fill


def del_cols(channel, sheet):
    """
    删除列
    """
    android_cols_to_delete = [7, 18, 19, 22]
    ios_cols_to_delete = [7, 18, 20, 23, 24, 28]
    flutter_cols_to_delete = [7, 18]
    server_cols_to_delete = [7, 9, 10, 12, 14, 16, 17, 21, 23, 25, 26, 29, 30, 31, 32, 33, 35]
    if channel == 'android':
        cols_to_delete = sorted(android_cols_to_delete, reverse=True)  # 确保从最大的列开始删除
        for col in cols_to_delete:
            sheet.delete_cols(col)
    elif channel == 'ios':
        cols_to_delete = sorted(ios_cols_to_delete, reverse=True)  # 确保从最大的列开始删除
        for col in cols_to_delete:
            sheet.delete_cols(col)
    elif channel == 'flutter':
        cols_to_delete = sorted(flutter_cols_to_delete, reverse=True)  # 确保从最大的列开始删除
        for col in cols_to_delete:
            sheet.delete_cols(col)
    else:
        cols_to_delete = sorted(server_cols_to_delete, reverse=True)  # 确保从最大的列开始删除
        for col in cols_to_delete:
            sheet.delete_cols(col)


def set_column_width(sheet, channel):
    """
    :param sheet: 继承sheet方法
    :param channel: 传入端名
    :return:设置列宽
    """
    if channel == 'android':
        for i in range(3, 27):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = 30
    elif channel == 'ios':
        for i in range(3, 24):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = 30


def get_head(file):
    """
    :param file: 传入下载文件
    :return: 返回文件表头
    """
    head = file.keys().tolist()
    return head


def change_head(file):
    """
    :param file:传入文件
    :return: 返回修改后的表头
    """
    head_list = get_head(file)
    lan_dic = language_dic()
    for i in range(len(head_list)):
        if head_list[i] in lan_dic:
            head_list[i] = lan_dic[head_list[i]]
    return head_list


def language_dic():
    lan_dic = {"ar": "ar：阿拉伯语🇸🇦", 'en': "en：英语🇬🇧", 'bn': "bn-IN：孟加拉语-印度🇧🇩",
               'bn-IN': "bn-IN：孟加拉语-印度🇧🇩",
               'cs': "cs：捷克语🇨🇿", 'de': "de：德语🇩🇪", 'es': "es：西班牙语🇪🇸", 'fr': "fr：法语🇫🇷", 'id': "id：印尼语🇮🇩",
               'in': "in：印尼语🇮🇩", "it": "it：意大利语🇮🇹", 'ja': "ja：日语🇯🇵", 'ko': "ko：韩语🇰🇷", 'ms': "ms：马来语🇲🇾",
               'pt-rBR': "pt-BR：葡萄牙语🇵🇹", 'pt-BR': "pt-BR：葡萄牙语🇵🇹", 'ru': "ru：俄语🇷🇺（老）",
               'ru-rRU': "ru-rRU：俄语🇷🇺", "ru-RU": "ru-rRU：俄语🇷🇺", 'sr': "sr：塞尔维亚语🇷🇸", 'th': "th：泰语🇹🇭",
               "tr": "tr：土耳其语🇹🇷（老）", 'tr-rTR': "tr-rTR：土耳其语🇹🇷（新）", 'tr-TR': "tr-TR：土耳其语🇹🇷",
               'ur-PK': "ur-rPK：ur-PK：乌尔都语🇵🇰", "ur-rPK": "ur-rPK：乌尔都语🇵🇰", 'vi': 'vi：越南语🇻🇳',
               'zh-rCN': "zh-rCN：中文🇨🇳", "zh-CN": "zh-CN：中文🇨🇳", 'zh-rTW': "zh-rTW：繁体中文🇨🇳",
               "zh-Hant": "zh-Hant:繁体中文🇨🇳"}
    return lan_dic


def change_filename(channel):
    """
    :param channel: 传入对应端名称
    :return: 文件名修改，并移动到指定文件夹
    """
    Dpath = os.path.expanduser(r'~\Downloads')
    # print(Dpath)
    times = time.strftime('%Y年%m月%d日 %H点-%M分-%S秒', time.localtime(time.time()))
    if channel == 'server':
        new_name = f"{absolute_path('data')}/{channel}_data/{times}language_{channel}.csv"
        en_file = find_file(Dpath, include_str='en', filter_strs=[".~"])
        shutil.move(en_file[0], new_name)
        print(f'{new_name}文件移动成功')
    else:
        new_name = f"{absolute_path('data')}/{channel}_data/{times}language_{channel}.xlsx"
        en_file = find_file(Dpath, include_str='en', filter_strs=[".~"])
        shutil.move(en_file[0], new_name)
        print(f'{new_name}文件移动成功')


# 文件移动
# def move_file(client):
#     change_filename(client)
#     data_path = f"D:/project/translate/data/{client}_data"
#     en_file = find_file(f'C:/Users/123123/Downloads', include_str=f'language_{client}', filter_strs=[".~"])
#     shutil.move(en_file[0], data_path)
#     sleep(3)


#
def insert_edit_cols(sheet):
    """
    :param sheet: 传入文件名
    :return: 插入编辑列
    """
    sheet.insert_cols(idx=1, amount=2)
    sheet["A2"] = "完成备注"
    sheet['B2'] = "编号"


def get_line_value(values, sheet):
    """
    :param values: 传入变化的内容变化内容
    :param sheet: 传入文件表
    :return: 向表格插入变化值
    """
    max1 = rows(language1)
    max2 = rows(language2)
    line = max1 - max2 + 4
    for i in range(len(values)):
        for k, v in values[i].items():
            for k1, v1 in v.items():
                line1 = letters_to_num(k1) + 3
                color_fill(sheet, line, line1)
                comment = Comment(str(v1), "hhhh")
                letters = number_to_letter(line1)
                str_num = str(line)
                nl = letters + str_num
                sheet[nl].comment = comment
        line += 1


# 转化数字
def letters_to_num(letter):
    """
    :param letter: 表格字母
    :return:返回对应的数字
    """
    result = ""
    for char in letter:
        if char.isdigit():
            result += char
    return int(result)


def number_to_letter(number):
    """
    :param number: 传入对应行的数
    :return: 数字转化为字母
    """
    if 1 <= number <= 26:
        return string.ascii_uppercase[number - 1]
    elif number > 26:
        return "A" + string.ascii_uppercase[number - 26 - 1]


def color_fill(sheet, row, column):
    """
    :param sheet: 传入的对应表格地址
    :param row: 对应的行
    :param column:
    :return:
    """
    fill = PatternFill('solid', fgColor='f8c600')
    sheet.cell(row, column).fill = fill


def del_file():
    """
    :return:清空result下文件
    """
    import os
    folder_path = fr"{absolute_path('result')}"
    file_list = os.listdir(folder_path)
    for file in file_list:
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            os.remove(file_path)

    print("已成功删除文件夹下的所有文件")


def absolute_path(data):
    """
    :return: 获取绝对路径
    """
    folder_name = f'{data}'
    folder_path = os.path.abspath(folder_name)
    return folder_path


def sort_excel_by_first_column_desc(file_path):
    """
    :param file_path:
    :return: 返回新的排序
    """
    # 读取 Excel 文件
    df = pd.read_excel(file_path)

    # 获取第一列列名
    first_col = df.columns[0]

    # 根据第一列倒序排序
    df_sorted = df.sort_values(by=first_col, ascending=False)

    # 保存回文件（可修改为另存为）
    df_sorted.to_excel(file_path, index=False)
    time.sleep(2)
    print(f"{file_path} 排序完成。")

